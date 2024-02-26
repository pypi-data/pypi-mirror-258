#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2022/10/21 10:43 AM
# @Author  : zhengyu.0985
# @FileName: crawler_weibo.py
# @Software: PyCharm

import requests
from requests import Response
import re
import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from bs4 import BeautifulSoup
from openpyxl import Workbook
import logging

logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s')  # logging.basicConfig函数对日志的输出格式及方式做相关配置
logger = logging.getLogger('rolling_king.jason.crawler.crawler_weibo')


# 微博热搜
class WeiBoCollection(object):

    headers = {
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
        "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36",
        "cookie": ""
    }

    def __init__(self, cookie_val):
        self.headers['cookie'] = cookie_val
        self.host_url = 'https://s.weibo.com/weibo'
        self.content_url = self.host_url
        self.wb = Workbook()

    # 获取微博热搜
    def get_hot_query_by_key(self, key: str) -> Response:
        hot_resp = requests.get(url="https://weibo.com/ajax/side/search?q="+key,
                                headers={
                                    "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36"}
                                )
        logger.info(f'微博热搜={hot_resp.json()}')
        return hot_resp

    # 微博内容
    def get_weibo_html_content_by_key(self, key: str) -> str:
        self.content_url = self.host_url+'?q=' + key + '&nodup=1'  # nodup=1代表查看微博全部结果
        content_resp = requests.get(url=self.content_url, headers=self.headers)
        print(content_resp.encoding)  # ISO-8859-1
        print(content_resp.apparent_encoding)  # GB2312
        # content_resp.encoding = content_resp.apparent_encoding
        # print(content_resp.content)  # bytes
        # print(content_resp.text)  # str
        return content_resp.text  # html_doc

    def get_total_page_num(self, html_doc: str = None) -> int:
        soup = BeautifulSoup(html_doc, "lxml")
        print(type(soup))
        # print(soup.find('a', class_='pagenum').text)
        ul_tag = soup.find('ul', attrs={'action-type': 'feed_list_page_morelist'})
        print(f'ul_tag={ul_tag}')
        page_num: int = len(ul_tag.find_all('li'))
        print('length=', page_num)
        return page_num

    def collect_func(self, curr_page: int) -> dict:
        print(f'current page = {curr_page}')
        curr_url = self.content_url + '&page=' + str(curr_page)
        print(f'current url = {curr_url}')
        curr_resp = requests.get(url=curr_url, headers=self.headers)
        curr_html_doc = curr_resp.text
        curr_soup = BeautifulSoup(curr_html_doc, "lxml")
        # from_results = curr_soup.find_all('div', class_='from')
        # print(len(from_results))
        results = curr_soup.find_all('p', class_='txt', attrs={'node-type': 'feed_list_content'})
        # results = curr_soup.find_all('p', class_='txt', attrs={'node-type': 'feed_list_content_full'})
        print(len(results))
        print(type(results))
        print(results)
        count: int = 0
        curr_dict = {
            'content': []
        }
        for item in results:
            count += 1
            print(type(item))
            print(item.name)  # p
            print(f"微博名={item['nick-name']}")  # 微博名
            print(f'微博内容={item.text.strip()}')  # 微博内容
            regex = re.compile(r'#.*?#')
            s = regex.search(item.text.strip())
            topic: str = ''
            if s is not None:
                print(f'话题={s.group()}')
                topic = s.group()
            curr_dict['content'].append({
                '微博名': item['nick-name'],
                '微博话题': topic,
                '微博内容': item.text.strip(),
            })
        print(f'--- 第{curr_page}页的{count}记录已获取 ---')
        curr_dict['count'] = count
        return curr_dict

    def save_weibo_content(self, page_num: int, key: str):
        thread_pool = ThreadPoolExecutor(page_num)
        thread_task_list = []
        for page in range(1, page_num+1):
            thread_task_list.append(thread_pool.submit(self.collect_func, page))

        print(self.wb.sheetnames)
        print(self.wb.active)
        ws = self.wb.active
        ws.title = key
        ws.cell(1, 1).value = '微博名'
        ws.cell(1, 2).value = '微博话题'
        ws.cell(1, 3).value = '微博内容'

        total_count = 0
        curr_row = 2
        for future in as_completed(thread_task_list):
            print(future.result())
            total_count += future.result()['count']
            # 存入Excel
            # 将一页的结果存入
            for dict_val in future.result()['content']:
                curr_col = 1
                ws.cell(curr_row, curr_col).value = dict_val['微博名']
                curr_col += 1
                ws.cell(curr_row, curr_col).value = dict_val['微博话题']
                curr_col += 1
                ws.cell(curr_row, curr_col).value = dict_val['微博内容']
                curr_row += 1
            # 一页的结果存完，从下一行存下一页的结果。
        print(f'{page_num}页，一共{total_count}条记录')

    def save_weibo_hot_query(self, hot_resp, key: str):
        ws = self.wb.create_sheet(title='热搜_' + key)
        if hot_resp.json()['ok'] == 1:
            hot_query_json_list = hot_resp.json()['data']['hotquery']
            if len(hot_query_json_list) > 0:
                key_list = hot_query_json_list[0].keys()
                curr_col = 1
                for col_head in key_list:
                    ws.cell(1, curr_col).value = col_head
                    curr_col += 1
                curr_row = 2
                for hot_query_json_item in hot_query_json_list:
                    curr_col = 1
                    for col_key in key_list:
                        ws.cell(curr_row, curr_col).value = hot_query_json_item[col_key]
                        curr_col += 1
                    curr_row += 1
            else:
                print(f'hot_query_json_list is empty.')
        else:
            print(f'hot_resp is not ok.')

    def save_excel_to_disk(self, file_name: str) -> None:
        self.wb.save(file_name)


if __name__ == '__main__':
    cookie_value: str = "XSRF-TOKEN=vH8eCkgP-JmRtN2Ia3VIZzNL; _s_tentry=weibo.com; Apache=8524959161901.953.1666270664916; SINAGLOBAL=8524959161901.953.1666270664916; ULV=1666270664920:1:1:1:8524959161901.953.1666270664916:; login_sid_t=b5127687703bbdcf584d351ad19bb4b4; cross_origin_proto=SSL; SSOLoginState=1666324094; SCF=ApUmMbNmgFup8JyPq2IgXMlCgCtSeadR43NF9Z6NG0KDyxJmqoy-q1BssnHP28j1ZKJlwOhyLRZzMNmw1cJ-FiM.; SUB=_2A25OUZ7RDeRhGedJ6VcV-SrLyDyIHXVtJvcZrDV8PUNbmtANLVr5kW9NVlLIFhGf5-a2Sp9qM7dSRByY1wlD_sSP; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9WW7wgizECTwskArQ2OMHFNw5JpX5KMhUgL.Fo2Neo-X1KBNe052dJLoIE-LxKnLB.-LB.xWi--4iKn0iK.pi--fi-z7iKysi--4iKn0iK.p; ALF=1698112000; WBPSESS=fbOmJTuMY3c-5Rw73SivynCCuNFzmQGVExuu7n6msq-AjXm4uN--xLuIUTml8RhJDN_nrrqPS1nQ2NIMyMdVyNKkaKtQladJWypSdM_rIwgLWcjOOCCCyt2nzPJT3IGPbG6yCmzbwCeOSpYz_m0h4g=="
    search_key = "南通"
    obj = WeiBoCollection(cookie_val=cookie_value)
    obj.save_weibo_content(obj.get_total_page_num(obj.get_weibo_html_content_by_key(key=search_key)), key=search_key)
    obj.save_weibo_hot_query(obj.get_hot_query_by_key(key=search_key), key=search_key)
    obj.save_excel_to_disk(file_name='WeiBo_'+datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')+'.xlsx')








