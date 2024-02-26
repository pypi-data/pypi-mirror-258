#!/usr/bin/python3
# -*- coding: UTF-8 -*-

from openpyxl import Workbook, load_workbook
from openpyxl.styles import *

# 爬取城市肯德基餐厅的位置信息 http://www.kfc.com.cn/kfccda/storelist/index.aspx

"""
抓包获取的数据
Request URL: http://www.kfc.com.cn/kfccda/ashx/GetStoreList.ashx?op=keyword
Request Method: POST
Status Code: 200 OK
Remote Address: 120.92.131.8:80
Referrer Policy: no-referrer-when-downgrade
"""

import requests
import json

url = 'http://www.kfc.com.cn/kfccda/ashx/GetStoreList.ashx?op=keyword'

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.20 Safari/537.36'
}

##############

city = input('input a city:')

data = {
    'cname': '',
    'pid': '',
    'keyword': city,
    'pageIndex': 0,
    'pageSize': 10,
}

response = requests.post(url=url, headers=headers, data=data)
print(response.json())

jsonVal = response.json()
listVal = jsonVal['Table']
print("列表长度={0}".format(len(listVal)))
for item in listVal:
    numb = item['rowcount']

data = {
    'cname': '',
    'pid': '',
    'keyword': city,
    'pageIndex': 1,
    'pageSize': numb,
}
response = requests.post(url=url, headers=headers, data=data)
jsonVal = response.json()
print(jsonVal)

cols = len(jsonVal['Table1'][0])
print("字段个数=", cols)
wb = Workbook()
# ws = wb.create_sheet("Record")
print(wb.sheetnames)
print(wb.active)
ws = wb.active
ws.title = "Records"
currRow = 1
currCol = 0
for key in jsonVal['Table1'][0].keys():
    print(key, end=",")
    currCol = currCol + 1
    ws.cell(currRow, currCol).value = key
print()
    
for jsonItem in jsonVal['Table1']:
    currRow = currRow + 1
    currCol = 0
    for val in jsonItem.values():
        print(val, end=',')
        currCol = currCol + 1
        ws.cell(currRow, currCol).value = val
    print()
    
# for jsonItem in jsonVal['Table1']:
    # print("---------------------")
    # for k, v in jsonItem.items():
    #     print(k, "=", v)

wb.save("Temp.xlsx")


