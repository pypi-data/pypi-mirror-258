import csv
import functools
import logging
import shutil
from collections import namedtuple
from configparser import ConfigParser
from math import floor
from pathlib import Path
from typing import Callable, Dict, Generator, Iterable, List, Set, Tuple, Union

import sob
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.worksheet import Worksheet

from . import model
from .config import (
    ADDITIONAL_QUALIFIER_NEEDED_CLASS_WORDS,
    ATTRIBUTE_NAMING_GUIDELINES_AND_ANALYSIS_REPORT_USAGE_HTML_PATH,
    CATALOG_JSON_PATH, CATALOG_XLSX_PATH, OUTPUT_XLSX_COLUMN_DIMENSIONS)
from .utilities import iter_column_names, sanitize_ini

# Defining the data structure for used catalog
UsedCatalog = Dict[str, Dict[str, List[str]]]

_catalog_document_lru_cache: Callable[
    [], Callable[..., Callable[..., model.CatalogDocument]]
] = functools.lru_cache  # type: ignore

_extra_catalog_lru_cache: Callable[
    [], Callable[..., Callable[..., Tuple[Tuple[str, ...], Tuple[str, ...]]]]
] = functools.lru_cache  # type: ignore


@_catalog_document_lru_cache()
def get_catalog_document(
    path: Union[str, Path] = CATALOG_JSON_PATH
) -> model.CatalogDocument:
    with open(path) as catalog_json_io:
        assert isinstance(catalog_json_io, sob.abc.Readable)
        return model.CatalogDocument(catalog_json_io)


@_extra_catalog_lru_cache()
def get_extra_catalog() -> Tuple[Tuple[str, ...], Tuple[str, ...]]:
    logging.info("Extracting catalog in anv.ini file")
    config = ConfigParser()
    config.read(Path().absolute().joinpath("anv.ini"))
    additional_class_word_abbreviations: Tuple = tuple()
    if (
        "additional-catalog" in config
        and "class-word-abbreviations" in config["additional-catalog"]
    ):
        additional_class_word_abbreviations = sanitize_ini(
            config["additional-catalog"]["class-word-abbreviations"]
        )
    additional_acronyms: Tuple = tuple()
    if (
        "additional-catalog" in config
        and "acronyms" in config["additional-catalog"]
    ):
        additional_acronyms = sanitize_ini(
            config["additional-catalog"]["acronyms"]
        )
    AdditionalAcronyms = namedtuple(
        "AdditionalAcronyms",
        "additional_class_word_abbreviations additional_acronyms",
    )
    return AdditionalAcronyms(
        additional_class_word_abbreviations, additional_acronyms
    )


def setup_workbook() -> Workbook:
    workbook = Workbook()
    default_sheet = workbook["Sheet"]
    workbook.remove(default_sheet)
    return workbook


def set_columns_width(worksheet: Worksheet, width: List[int]) -> None:
    bold = Font(bold=True)
    dim_holder: DimensionHolder = DimensionHolder(worksheet=worksheet)
    for col in range(worksheet.min_column, worksheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            worksheet, min=col, max=col, width=width[col - 1]
        )
        cell = worksheet.cell(row=1, column=col)
        cell.font = bold
    worksheet.column_dimensions = dim_holder


class AttributeNameValidator:
    def __init__(
        self,
        column_names_file_path: Path,
        catalog_document_path: Union[str, Path] = CATALOG_JSON_PATH,
        extra_class_word_abbreviations: Tuple = (),
        extra_acronyms: Tuple = (),
        write_to_text_files: bool = True,
    ) -> None:
        self.column_names_file_path = column_names_file_path
        self.catalog_document_path = catalog_document_path
        self.catalog: model.CatalogDocument = get_catalog_document(
            catalog_document_path
        )
        self.update_catalog(extra_class_word_abbreviations, extra_acronyms)
        self.write_to_text_files = write_to_text_files
        self.entity_name = self.column_names_file_path.parts[-1].split(".")[0]
        self.class_word_analysis_report: model.ClassWordAnalysisReport = (
            model.ClassWordAnalysisReport()
        )
        self.full_words_used: Set[str] = set()
        self.catalog_used: UsedCatalog = dict()
        self.header_flag: bool = True
        self.workbook: Workbook = setup_workbook()
        self.reports_directory: Path = Path().absolute().joinpath("reports")
        self.reports_directory.mkdir(parents=True, exist_ok=True)
        self.entity_reports_directory: Path = self.reports_directory.joinpath(
            self.entity_name
        )

    def update_catalog(
        self,
        extra_class_word_abbreviations: Iterable[str],
        extra_acronyms: Iterable[str],
    ) -> None:
        """
        updating class word and acronym catalog with the custom list of extra
         words provided by anv.ini file
        """

        assert self.catalog.acronyms
        assert self.catalog.class_words
        assert self.catalog.class_word_abbreviations
        # updating class word dictionary with list of extra acronyms
        for class_word in extra_class_word_abbreviations:
            if (
                class_word not in self.catalog.class_words
                and class_word not in self.catalog.class_word_abbreviations
            ):
                self.catalog.class_word_abbreviations[class_word] = (
                    model.ClassWordAbbreviation(
                        abbreviation=class_word,
                        class_word=class_word,
                    )
                )
                self.catalog.class_words[class_word] = dict()
        # updating acronym dictionary with list of extra acronyms
        for acronym in extra_acronyms:
            if acronym not in self.catalog.acronyms:
                self.catalog.acronyms.setdefault(acronym, [acronym])

    def analyze_entity(self) -> None:
        """
        This function does column name analysis based on enterprise
        column naming guidelines on all column names of a given entity.
        """
        logging.info(f"Analyzing entity {self.entity_name}")
        column_names: Generator = iter_column_names(
            self.column_names_file_path
        )
        for column_name in column_names:
            # skip the analysis on header, if it exists
            if self.header_flag and (
                column_name == "name"
                or column_name == "NAME"
                or column_name == "COLUMN_NAME"
            ):
                self.header_flag = False
                continue

            self._analyze_column_name(column_name)

    def _analyze_column_name(self, column_name: str) -> None:
        """
        This function does analysis based on column naming
        guidelines on a given column name.
        """
        logging.info(f"analyzing {column_name}")
        self.class_word_analysis_report.setdefault(
            column_name, model.ColumnNameClassWordAnalysis()
        )
        self._check_class_word_column_naming(column_name)
        self._analyze_column_name_except_last_word(column_name)
        self._analyze_column_name_last_word(column_name)

    def _check_class_word_column_naming(self, column_name: str) -> None:
        """
        This function checks if a class word or its abbreviation itself is
        used as a whole name in naming a column.
        """
        assert self.catalog.class_words
        assert self.catalog.class_word_abbreviations
        if (
            column_name in self.catalog.class_word_abbreviations
            or column_name in self.catalog.class_words
        ):
            self.class_word_analysis_report[column_name].append(
                model.ColumnNameClassWordAnalysisUnit(
                    column_name=column_name,
                    word=column_name,
                    analysis="CLASS WORD IS USED AS COLUMN NAME",
                    class_word_rules_followed="NO",
                )
            )

    def _analyze_column_name_except_last_word(self, column_name: str) -> None:
        """
        This function does analysis on all words except last word

        1. Documents the approved short-forms used, their approved
        expansions(class words and acronyms), column name in which they appear
         to generate a report to be looked up by the end users, to make sure
         abbreviations were only used for approved ones.
         For example: if _STD_ is used by developers of object to mean
         _STANDARD_, not for "SEASON TO DATE" , an approved Acronym, then it
         is in violation of enterprise naming guidelines
        2. Documents the usage of class words in the middle of the column name
        which **may violate naming guidelines, if used w/o another class word
        postfix.
        3. Collect all words used in the naming of columns, which are outside
        approved abbreviations, for LONG_FORM_WORD report.
        This report will be used for look up by end users of the
        package, to quickly identify, if there are any shortened or composed
        words or acronyms, that are used in violation.
        """
        assert self.catalog.acronyms
        assert self.catalog.aggregates
        assert self.catalog.class_words
        assert self.catalog.class_word_abbreviations
        column_name_words = column_name.split("_")
        for index, word in enumerate(column_name_words[:-1]):
            word = word.upper()
            if (
                word in self.catalog.acronyms
                or word in self.catalog.class_word_abbreviations
            ):
                if word in self.catalog.acronyms:
                    self._update_catalog_used(word, column_name)
                if word in self.catalog.class_word_abbreviations:
                    self.class_word_analysis_report[column_name].append(
                        model.ColumnNameClassWordAnalysisUnit(
                            column_name=column_name,
                            word=word,
                            analysis="ABBREVIATED CLASS WORD IS USED IN THE "
                            "MIDDLE",
                            class_word_rules_followed="NO",
                            when_to_use=self.catalog.class_word_abbreviations[
                                word
                            ].when_to_use,
                        )
                    )
            elif word in self.catalog.class_words:
                # assumption: a class word that appears outside the last
                # quarter of name is probably not a class word,
                # rather a descriptor of the column
                if index >= floor(len(column_name_words) * 0.75):
                    self.class_word_analysis_report[column_name].append(
                        model.ColumnNameClassWordAnalysisUnit(
                            column_name=column_name,
                            word=word,
                            analysis="FULL CLASS WORD IS USED IN THE MIDDLE",
                            class_word_rules_followed="NO",
                        )
                    )
            elif word in self.catalog.aggregates:
                self.class_word_analysis_report[column_name].append(
                    model.ColumnNameClassWordAnalysisUnit(
                        column_name=column_name,
                        word=word,
                        analysis="AGGREGATE ACRONYM IS USED IN THE MIDDLE",
                        class_word_rules_followed="NO",
                        when_to_use=self.catalog.aggregates[word].when_to_use,
                    )
                )
            else:
                self.full_words_used.add(word)

    def _set_catalog_used_entry(self, word: str) -> None:
        if word not in self.catalog_used:
            self.catalog_used.setdefault(word, dict())
            self.catalog_used[word]["column_names"] = []
            self.catalog_used[word]["usages"] = []

    def _update_catalog_used(self, word: str, column_name: str) -> None:
        assert self.catalog.acronyms
        assert self.catalog.class_word_abbreviations

        self._set_catalog_used_entry(word)
        if column_name not in self.catalog_used[word]["column_names"]:
            self.catalog_used[word]["column_names"].append(column_name)
        if word in self.catalog.acronyms:
            if (
                self.catalog.acronyms[word][0]
                not in self.catalog_used[word]["usages"]
            ):
                self.catalog_used[word]["usages"].extend(
                    self.catalog.acronyms[word]
                )
        elif word in self.catalog.class_word_abbreviations:
            if (
                self.catalog.class_word_abbreviations[word].class_word
                not in self.catalog_used[word]["usages"]
            ):
                self.catalog_used[word]["usages"].append(
                    self.catalog.class_word_abbreviations[word].class_word
                )

    def _analyze_column_name_last_word(  # noqa:  C901
        self, column_name: str
    ) -> None:
        """
        This function checks if last word in the name of column is used as per
        enterprise naming guidelines.
        Also checks for class word usage in prior words of the column name, if
        they may have followed enterprise naming guidelines, as class word
        abbreviation usage in the middle could be warranted for some word
        """
        last_word = column_name.split("_")[-1].upper()
        assert self.catalog.acronyms
        assert self.catalog.aggregates
        assert self.catalog.class_words
        assert self.catalog.class_word_abbreviations
        # if last word is one of the following, then it may need to followed
        # by unit of its measure i.e AMT_USD, WT_KG, TM_DAYS etc
        if last_word in ADDITIONAL_QUALIFIER_NEEDED_CLASS_WORDS:
            self.class_word_analysis_report[column_name].append(
                model.ColumnNameClassWordAnalysisUnit(
                    column_name=column_name,
                    word=last_word,
                    analysis="UNIT SPECIFIC CLASS WORD MAY BE NEEDED AT THE "
                    "END",
                    additional_notes=ADDITIONAL_QUALIFIER_NEEDED_CLASS_WORDS[
                        last_word
                    ],
                    class_word_rules_followed="MAY BE",
                    when_to_use=self.catalog.class_word_abbreviations[
                        last_word
                    ].when_to_use,
                )
            )
        elif (
            last_word in self.catalog.class_word_abbreviations
            or last_word in self.catalog.aggregates
        ):
            if last_word in self.catalog.class_word_abbreviations:
                self.class_word_analysis_report[column_name].append(
                    model.ColumnNameClassWordAnalysisUnit(
                        column_name=column_name,
                        word=last_word,
                        analysis="ABBREVIATED CLASS WORD IS USED AT THE END",
                        class_word_rules_followed="YES",
                        when_to_use=self.catalog.class_word_abbreviations[
                            last_word
                        ].when_to_use,
                    )
                )
            else:
                self.class_word_analysis_report[column_name].append(
                    model.ColumnNameClassWordAnalysisUnit(
                        column_name=column_name,
                        word=last_word,
                        analysis="AGGREGATE ACRONYM IS USED AT THE END",
                        class_word_rules_followed="YES",
                        when_to_use=self.catalog.aggregates[
                            last_word
                        ].when_to_use,
                    )
                )

            for word_analysis in self.class_word_analysis_report[column_name]:
                if word_analysis.analysis in (
                    "ABBREVIATED CLASS WORD IS USED IN THE MIDDLE",
                    "FULL CLASS WORD IS USED IN THE MIDDLE",
                ):
                    word_analysis.class_word_rules_followed = "YES"
        else:
            if last_word in self.catalog.class_words:
                for word_analysis in self.class_word_analysis_report[
                    column_name
                ]:
                    if word_analysis.analysis in (
                        "FULL CLASS WORD IS USED IN THE MIDDLE",
                        "ABBREVIATED CLASS WORD IS USED IN THE MIDDLE",
                    ):
                        word_analysis.class_word_rules_followed = "MAY BE"
                self.class_word_analysis_report[column_name].append(
                    model.ColumnNameClassWordAnalysisUnit(
                        column_name=column_name,
                        word=last_word,
                        analysis="FULL CLASS WORD IS USED AT THE END",
                        class_word_rules_followed="NO",
                    )
                )
            elif last_word in self.catalog.acronyms:
                self._update_catalog_used(last_word, column_name)
                self.class_word_analysis_report[column_name].append(
                    model.ColumnNameClassWordAnalysisUnit(
                        column_name=column_name,
                        word=last_word,
                        analysis="APPROVED ACRONYM IS USED AT THE END",
                        class_word_rules_followed="NO",
                    )
                )
            else:
                self.class_word_analysis_report[column_name].append(
                    model.ColumnNameClassWordAnalysisUnit(
                        column_name=column_name,
                        word=last_word,
                        analysis="GENERIC WORD USED AT THE END",
                        class_word_rules_followed="NO",
                    )
                )
        class_word_exists: bool = False
        analysis_unit: model.ColumnNameClassWordAnalysisUnit
        for analysis_unit in self.class_word_analysis_report[column_name]:
            assert analysis_unit.analysis
            if (
                analysis_unit.analysis
                and "CLASS WORD" in analysis_unit.analysis
            ):
                class_word_exists = True
        if not class_word_exists:
            self.class_word_analysis_report[column_name].append(
                model.ColumnNameClassWordAnalysisUnit(
                    column_name=column_name,
                    word="",
                    analysis="NO CLASS WORD IS USED IN THE NAME",
                    class_word_rules_followed="NO",
                )
            )

    def save_reports(self) -> None:
        """
        Independent of where the input file is from, create a "reports" folder
         in the current working directory of cli, and save entity specific
         reports in directory immediately below the "reports" directory.
        """

        logging.info(f"saving reports of {self.entity_name}")

        shutil.copy(
            ATTRIBUTE_NAMING_GUIDELINES_AND_ANALYSIS_REPORT_USAGE_HTML_PATH,
            self.reports_directory.joinpath(
                "ATTRIBUTE_NAMING_GUIDELINES_AND_ANALYSIS_REPORT_USAGE.html"
            ),
        )
        shutil.copy(
            CATALOG_XLSX_PATH, self.reports_directory.joinpath("CATALOG.xlsx")
        )

        self.save_to_xlsx(
            path=self.reports_directory.joinpath(
                f"{self.entity_name}_REPORT.xlsx"
            ),
        )
        # with the decision to create xlsx reports for end user,
        # for the purposes of development, changes to csv
        # reports are easier to track, as we change catalog,
        # naming rules.
        if self.write_to_text_files:
            # create directory per entity to store the csv files in
            self.entity_reports_directory.mkdir(parents=True, exist_ok=True)
            self.save_csv_reports()

    def save_to_xlsx(
        self,
        path: Path,
    ) -> None:
        self.load_class_word_analysis()
        self.load_full_words_used()
        self.load_catalog_used()

        # Save the xlsx file
        self.workbook.save(f"{path}")

    def load_class_word_analysis(self) -> None:
        worksheet = self.workbook.create_sheet("CLASS WORD ANALYSIS")
        (
            worksheet["A1"],
            worksheet["B1"],
            worksheet["C1"],
            worksheet["D1"],
            worksheet["E1"],
            # worksheet["F1"],
        ) = (
            "COLUMN",
            "CLASS WORD",
            "ANALYSIS",
            "CLASS WORD RULES FOLLOWED?",
            "ADDITIONAL NOTES",
            # "WHEN TO USE",
        )

        for column_name in self.class_word_analysis_report:
            for (
                column_name_class_word_analysis_unit
            ) in self.class_word_analysis_report[column_name]:
                worksheet.append(
                    [
                        column_name_class_word_analysis_unit.column_name,
                        column_name_class_word_analysis_unit.word,
                        column_name_class_word_analysis_unit.analysis,
                        column_name_class_word_analysis_unit.class_word_rules_followed,  # noqa
                        column_name_class_word_analysis_unit.additional_notes,
                        # column_name_class_word_analysis_unit.when_to_use,
                    ]
                )

        set_columns_width(
            worksheet, OUTPUT_XLSX_COLUMN_DIMENSIONS["CLASS_WORD_ANALYSIS"]
        )

    def load_full_words_used(self) -> None:
        worksheet: Worksheet = self.workbook.create_sheet("FULL WORDS USED")
        worksheet["A1"] = "WORD"
        full_words_used: List = list(self.full_words_used)
        full_words_used.sort()
        for word in full_words_used:
            worksheet.append([word])

        set_columns_width(
            worksheet, OUTPUT_XLSX_COLUMN_DIMENSIONS["FULL_WORDS_USED"]
        )

    def load_catalog_used(self) -> None:
        worksheet: Worksheet = self.workbook.create_sheet(
            "APPROVED ACRONYMS USED"
        )
        worksheet["A1"], worksheet["B1"], worksheet["C1"] = (
            "ABBREVIATION",
            "COLUMN",
            "USAGES ALLOWED",
        )
        for abbreviation in self.catalog_used:
            for column_name in self.catalog_used[abbreviation]["column_names"]:
                worksheet.append(
                    [
                        abbreviation,
                        column_name,
                        ",".join(self.catalog_used[abbreviation]["usages"]),
                    ]
                )

        set_columns_width(
            worksheet,
            OUTPUT_XLSX_COLUMN_DIMENSIONS["APPROVED_ACRONYMS_USED"],
        )

    def save_csv_reports(self) -> None:
        self.save_class_word_analysis_csv()
        self.save_approved_acronyms_used_csv()
        self.save_full_word_csv()

    def save_class_word_analysis_csv(self) -> None:
        with open(
            self.entity_reports_directory.joinpath("CLASS_WORD_ANALYSIS.csv"),
            "w",
        ) as f:
            csv_writer = csv.writer(f)
            csv_writer.writerow(
                [
                    "COLUMN",
                    "CLASS WORD",
                    "ANALYSIS",
                    "CLASS WORD RULES FOLLOWED?",
                    "ADDITIONAL NOTES",
                ]
            )
            for column_name in self.class_word_analysis_report:
                for (
                    column_name_class_word_analysis_unit
                ) in self.class_word_analysis_report[column_name]:
                    csv_writer.writerow(
                        [
                            column_name_class_word_analysis_unit.column_name,
                            column_name_class_word_analysis_unit.word,
                            column_name_class_word_analysis_unit.analysis,
                            column_name_class_word_analysis_unit.class_word_rules_followed,  # noqa
                            column_name_class_word_analysis_unit.additional_notes,  # noqa
                        ]
                    )

    def save_approved_acronyms_used_csv(self) -> None:
        with open(
            self.entity_reports_directory.joinpath(
                "APPROVED_ACRONYMS_USED.csv"
            ),
            "w",
        ) as f:
            csv_writer = csv.writer(f)
            csv_writer.writerow(
                [
                    "ACRONYM",
                    "COLUMN",
                    "USAGES ALLOWED",
                ]
            )
            for abbreviation in self.catalog_used:
                for column_name in self.catalog_used[abbreviation][
                    "column_names"
                ]:
                    csv_writer.writerow(
                        [
                            abbreviation,
                            column_name,
                            ",".join(
                                self.catalog_used[abbreviation]["usages"]
                            ),
                        ]
                    )

    def save_full_word_csv(self) -> None:
        with open(
            self.entity_reports_directory.joinpath("FULL_WORDS_USED.csv"), "w"
        ) as f:
            """
            Sorting, so that every run doesn't create a new order of words.
            Makes it easier to track changes with addition of new class
            words/ acronyms/ rules
            """
            full_words_used: List = list(self.full_words_used)
            full_words_used.sort()
            csv_writer = csv.writer(f)
            csv_writer.writerow(["FULL_WORD"])
            for word in full_words_used:
                csv_writer.writerow([word])
