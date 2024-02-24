from __future__ import annotations
import logging
from pathlib import Path
from typing import Any
from .inout.utils import Row
from . import files
from openpyxl import load_workbook, Workbook, DEFUSEDXML
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableColumn, TableFormula, TableStyleInfo
from openpyxl.worksheet.formula import DataTableFormula, ArrayFormula
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.cell.cell import Cell
from openpyxl.styles.differential import DifferentialStyle, DifferentialStyleList
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import range_boundaries, get_column_letter

logger = logging.getLogger(__name__)


class ExcelWorkbook:
    _workbook_cache: dict[str,ExcelWorkbook] = {}
    """ Workbooks per canonical path """
    
    _defusedxml_alert_emitted = False

    def __init__(self, path: str|Path):        
        if not DEFUSEDXML and not self.__class__._defusedxml_alert_emitted:
            logger.warning("By default openpyxl does not guard against quadratic blowup or billion laughs xml attacks. To guard against these attacks install defusedxml.")
            self.__class__._defusedxml_alert_emitted = True

        self.path = Path(path) if not isinstance(path, Path) else path

        if files.exists(self.path):
            logger.info("load excel workbook %s", self.path)
            with files.open(self.path, 'rb') as fp:
                self.pyxl_workbook: Workbook = load_workbook(fp) #NOTE: using keep_vba=True causes issues, example on openpyxl 3.1.2: load and save samples/empty-table.xlsx without modification corrupts the file
            self._create_next_table_in_active_sheet = False
        else:
            logger.info("create excel workbook for %s", self.path)
            self.pyxl_workbook = Workbook()
            self._create_next_table_in_active_sheet = True

        self.tables: dict[str,ExcelTable] = {}
        self.needs_save = False


    @classmethod
    def get_cached(cls, path: str|Path) -> ExcelWorkbook:
        path = Path(path) if not isinstance(path, Path) else path
        canonical_path = str(path.resolve())
        return cls._workbook_cache[canonical_path]


    @classmethod
    def get_or_create_cached(cls, path: str|Path) -> ExcelWorkbook:
        path = Path(path) if not isinstance(path, Path) else path
        canonical_path = str(path.resolve())
        if not canonical_path in cls._workbook_cache:
            cls._workbook_cache[canonical_path] = ExcelWorkbook(path)
        return cls._workbook_cache[canonical_path]


    @classmethod
    def close_all_cached(cls):
        for workbook in cls._workbook_cache.values():
            workbook.close()


    def close(self):
        if not self.needs_save:
            logger.info("excel workbook %s not modified", self.path)
            return
        
        for table in self.tables.values():
            table.redefine()
        
        logger.info("save excel workbook %s", self.path)
        with files.open(self.path, 'wb') as fp:
            self.pyxl_workbook.save(fp)

        self.needs_save = False
        

    def get_table(self, name: str, default = '__raise__') -> ExcelTable:
        if name in self.tables:
            return self.tables[name]
        
        for sheet_name in self.pyxl_workbook.sheetnames:
            pyxl_worksheet: Worksheet = self.pyxl_workbook[sheet_name]
            if name in pyxl_worksheet.tables:
                pyxl_table = pyxl_worksheet.tables[name]
                self.tables[name] = ExcelTable(pyxl_table, pyxl_worksheet, self)
                return self.tables[name]
            
        if default == '__raise__':
            raise KeyError(f"no table found with name \"{name}\" in workbook \"{self.path}\"")
        else:
            return default
        

    def create_table(self, name: str, no_headers: bool = False) -> ExcelTable:
        for sheet_name in self.pyxl_workbook.sheetnames:
            pyxl_worksheet: Worksheet = self.pyxl_workbook[sheet_name]
            if name in pyxl_worksheet.tables:
                raise ValueError(f"table {name} already exist")
    
        self.needs_save = True
        
        if self._create_next_table_in_active_sheet:
            pyxl_worksheet: Worksheet = self.pyxl_workbook.active
            pyxl_worksheet.title = name
            self._create_next_table_in_active_sheet = False
        else:
            pyxl_worksheet: Worksheet = self.pyxl_workbook.create_sheet(title=name)

        self.tables[name] = ExcelTable(name, pyxl_worksheet, self, no_headers=no_headers)
        return self.tables[name]

        
    def get_or_create_table(self, name: str, no_headers: bool = False) -> ExcelTable:
        table = self.get_table(name, default=None)
        if table is None:
            table = self.create_table(name, no_headers)
        return table
    
    # defined names

    def get_global_named_values(self, name: str) -> list:
        defn = self.pyxl_workbook.defined_names[name]
        values = []
        
        for title, coord in defn.destinations:
            worksheet = self.pyxl_workbook[title]
            cell = worksheet[coord]
            value = cell.value
            values.append(value)

        return values

    def get_global_named_value(self, name: str) -> Any:
        values = self.get_global_named_values(name)
        
        if len(values) > 1:
            raise ValueError(f"more than one cell")
        elif len(values) == 0:
            raise ValueError(f"global name not found")
        else:
            return values[0]



class ExcelTable:
    min_row_index: int
    """ 0-base index of the first data row (excluding headers). """
    
    min_col_index: int
    """ 0-base index of the first column. """

    row_count: int
    """ Number of data rows (excluding headers). """

    col_count: int
    """ Number of columns. """

    def __init__(self, pyxl_table: Table|str, pyxl_worksheet: Worksheet, workbook: ExcelWorkbook, no_headers: bool = None):
        self.workbook = workbook
        self.pyxl_worksheet = pyxl_worksheet        

        if isinstance(pyxl_table, str):
            self.pyxl_table: Table = None
            self.name: str = pyxl_table
            if no_headers:
                self.has_headers = False
                self.min_row_index = 0
            else:
                self.has_headers = True
                self.min_row_index = 1
            self.min_col_index = 0
            self.row_count = 0
            self.column_names: list[str] = []

        elif isinstance(pyxl_table, Table):
            self.pyxl_table: Table = pyxl_table
            self.name: str = pyxl_table.name
            
            min_col, min_row, max_col, max_row = range_boundaries(pyxl_table.ref)

            # NOTE: range_boundaries returns 1-base indices
            self.min_col_index = min_col - 1
            if pyxl_table.headerRowCount == 0:
                self.has_headers = False
                self.min_row_index = min_row - 1
            elif pyxl_table.headerRowCount == 1:
                self.has_headers = True
                self.min_row_index = min_row
            else:
                raise ValueError(f'invalid headerRowCount: {pyxl_table.headerRowCount}')
            
            self.row_count = max_row - self.min_row_index
            col_count = max_col - self.min_col_index
        
            self.column_names: list[str] = list(self.pyxl_table.column_names)
            # NOTE: if no headers, default names are returned, example in French: ['Colonne1', 'Colonne2']
            if len(self.column_names) != col_count:
                raise ValueError(f'invalid column_names length ({len(self.column_names)}, expected {col_count}): {self.column_names}')

            if self.row_count == 1 and self.is_row_empty(0):
                self.row_count = 0
        else:
            raise ValueError(f"invalid type for pyxl_table: {type(pyxl_table).__name__}")

        self._column_formats: list[dict[str,Any]] = None
    
    @property
    def col_count(self) -> int:
        return len(self.column_names)
    
    @property
    def ref(self) -> str:
        if self.col_count == 0:
            raise ValueError(f"cannot get table ref: table does not contain any column")
        if self.row_count == 0:
            raise ValueError(f"cannot get table ref: table does not contain any row")
                
        return f"{get_column_letter(self.min_col_index + 1)}{self.min_row_index - (1 if self.has_headers else 0) + 1}:{get_column_letter(self.min_col_index + self.col_count)}{self.min_row_index + self.row_count}"


    def get_row(self, index: int, readonly: bool = False) -> Row:
        """
        Get row at the given 0-base index.
        """
        if index == -1:
            if self.row_count == 0:
                raise ValueError(f"cannot get last row: table does not contain any row")
            index = self.row_count - 1
        elif index < 0:
            raise ValueError(f"invalid row index: {index}")
        
        if index >= self.row_count:
            raise ValueError(f"cannot get row at index {index}: table contains {self.row_count} rows")

        return Row(
            get=lambda col_index: self.get_value(index, col_index),
            set=None if readonly else (lambda col_index, value: self.set_value(index, col_index, value)),
            headers=self.column_names,
            index=index,
        )
    

    def is_row_empty(self, row_index: int):
        for col_index in range(0, self.col_count):
            cell = self.pyxl_worksheet.cell(self.min_row_index + 1 + row_index, self.min_col_index + 1 + col_index)
            if cell.value is not None:
                return False
            
        return True
    

    def is_col_empty(self, col_index: int):
        for row_index in range(0, self.row_count):
            cell = self.pyxl_worksheet.cell(self.min_row_index + 1 + row_index, self.min_col_index + 1 + col_index)
            if cell.value is not None:
                return False
            
        return True
    

    def insert_row(self) -> Row:
        self.workbook.needs_save = True
        row_index = self.row_count

        if not self.is_row_empty(row_index):
            raise ValueError(f'cannot insert row: row at index {row_index} is not empty')

        self.row_count += 1

        # erase old styles and apply column format
        for col_index in range(0, self.col_count):
            self.erase_cell(row_index, col_index)
            
        return self.get_row(row_index)
    

    def insert_col(self, name: str):
        self.workbook.needs_save = True
        if not name:
            raise ValueError(f'name cannot be empty')
        if name in self.column_names:
            raise ValueError(f'column name already used: {name}')
            
        col_index = self.col_count

        if not self.is_col_empty(col_index):
            raise ValueError(f'cannot insert column: column {col_index} is not empty')

        self.column_names.append(name) # implies self.col_count += 1

        if self.has_headers:
            cell = self.pyxl_worksheet.cell(self.min_row_index, self.min_col_index + 1 + col_index)
            #logger.debug('set column name to cell %s%s: %s', cell.column_letter, cell.row, name)
            cell.value = name

        # erase old styles and apply column format
        for row_index in range(0, self.row_count):
            self.erase_cell(row_index, col_index)
            

    def truncate(self):
        self.workbook.needs_save = True
        prev_row_count = self.row_count
        self.row_count = 0

        for row_index in range(0, prev_row_count):
            for col_index in range(0, self.col_count):
                self.erase_cell(row_index, col_index, allow_outside=True)


    def redefine(self):
        self.workbook.needs_save = True

        # table cannot be empty (must have at least one blank row)
        if self.row_count == 0:
            self.insert_row()

        new_ref = self.ref
        if self.pyxl_table is not None and new_ref == self.pyxl_table.ref:
            return
        
        logger.debug("define table %s: %s => %s", self.name, self.pyxl_table.ref if self.pyxl_table is not None else None, new_ref)

        newcolumns = []

        for i in range(0, self.col_count):
            if self.has_headers:
                name = self.pyxl_worksheet.cell(self.min_row_index, self.min_col_index + 1 + i).value
            else:
                name = self.column_names[i] if i < len(self.column_names) else None
            newcolumn = TableColumn(id=i+1, name=name)
            newcolumns.append(newcolumn)

            if self.pyxl_table is not None and i < len(self.pyxl_table.tableColumns):
                prevcolumn: TableColumn = self.pyxl_table.tableColumns[i]
                newcolumn.dataCellStyle = prevcolumn.dataCellStyle
                newcolumn.dataDxfId = prevcolumn.dataDxfId # refers to workbook._differential_styles
                newcolumn.calculatedColumnFormula = prevcolumn.calculatedColumnFormula


        newtable_kwargs = {
            'name': self.name,
            'displayName': self.name,
            'ref': new_ref,
            'tableColumns': newcolumns,
            'headerRowCount': 1 if self.has_headers else 0,
        }

        if self.pyxl_table is not None:
            newtable_kwargs['autoFilter'] = self.pyxl_table.autoFilter
            newtable_kwargs['sortState'] = self.pyxl_table.sortState
            newtable_kwargs['tableStyleInfo'] = self.pyxl_table.tableStyleInfo
        else:
            newtable_kwargs['autoFilter'] = AutoFilter()
            newtable_kwargs['tableStyleInfo'] = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

        newtable = Table(**newtable_kwargs)

        self.pyxl_table = newtable
        
        if self.name in self.pyxl_worksheet.tables:
            del self.pyxl_worksheet.tables[self.name]
        self.pyxl_worksheet.add_table(newtable)


    def get_value(self, row_index: int, col_index: int) -> Any:
        self._check_indexes(row_index, col_index)
        cell = self.pyxl_worksheet.cell(self.min_row_index + 1 + row_index, self.min_col_index + 1 + col_index)
        value = cell.value
        # TODO: formatting?
        return value
    

    def set_value(self, row_index: int, col_index: int, value: Any):
        """
        Set the value of the cell located at the given 0-base indices, and apply the default formatting and formulas of the corresponding table column.

        Any value (including `None`) overrides default column formulas. If you want to use the default column formula, use `erase_cell` method instead.
        """
        self._check_indexes(row_index, col_index)
        self.workbook.needs_save = True
        cell = self.pyxl_worksheet.cell(self.min_row_index + 1 + row_index, self.min_col_index + 1 + col_index)

        self._apply_column_format(cell)

        try:
            #logger.debug('set value to cell %s%s: %s', cell.column_letter, cell.row, value)
            cell.value = value
        except ValueError as err:
            if str(err).startswith('Cannot convert'):
                cell.value = str(value)
            else:
                raise
        

    def _check_indexes(self, row_index: int, col_index: int):
        if row_index == -1:
            row_index = self.row_count - 1
        if col_index == -1:
            col_index = self.col_count - 1

        if row_index < 0 or row_index >= self.row_count:
            raise ValueError(f"invalid row index: {row_index} (row count: {self.row_count})")
        if col_index < 0 or col_index >= self.col_count:
            raise ValueError(f"invalid row index: {col_index} (row count: {self.col_count})")


    def erase_cell(self, row_index: int, col_index: int, allow_outside: bool = False):
        """
        Erase the value of the cell located at the given 0-base indices, and apply the default formatting and formulas of the corresponding table column.
        
        If `allow_outside` is set, the cell may be located outside of the table. In this case, no formatting or formula is applied.
        """
        if not allow_outside:
            self._check_indexes(row_index, col_index)

        self.workbook.needs_save = True
        cell = self.pyxl_worksheet.cell(self.min_row_index + 1 + row_index, self.min_col_index + 1 + col_index)
        cell.style = 'Normal'
        #logger.debug('erase cell %s%s', cell.column_letter, cell.row)
        cell.value = None

        if not allow_outside or (row_index < self.row_count and col_index < self.col_count):
            self._apply_column_format(cell)
    

    def _apply_column_format(self, cell: Cell):
        if self._column_formats is None:
            self._column_formats = self._build_column_formats()
        
        index = (cell.col_idx - 1) - self.min_col_index
        if index >= len(self._column_formats):
            return
        
        fmt = self._column_formats[index]

        if 'formula' in fmt:
            formula = fmt['formula']
            if isinstance(formula, ArrayFormula):
                pass # TODO: not supported yet
            else:
                #logger.debug('apply formula to cell %s%s: %s', cell.column_letter, cell.row, formula)
                cell.value = formula

        if 'style' in fmt:
            cell.style = fmt['style']

        for fmt_key, fmt_value in fmt.items():
            if fmt_key in ['formula', 'style']:
                continue
            setattr(cell, fmt_key, fmt_value)


    def _build_column_formats(self) -> list[dict[str,Any]]:
        if not self.pyxl_table:
            return []
                
        column: TableColumn
        fmt_list = []
        for index, column in enumerate(self.pyxl_table.tableColumns):
            fmt: dict[str,Any] = {}
            fmt_list.append(fmt)

            # Read dataCellStyle
            if column.dataCellStyle:
                fmt['style'] = column.dataCellStyle
            
            # Read dxf
            if column.dataDxfId is not None:
                dxf: DifferentialStyle = self.workbook.pyxl_workbook._differential_styles[column.dataDxfId]

                if dxf.numFmt:
                    fmt['number_format'] = dxf.numFmt.formatCode
                else:
                    if not 'style' in fmt:
                        fmt['number_format'] = self._DEFAULT_NUMBER_FORMAT

                fmt['alignment'] = dxf.alignment if dxf.alignment else self._DEFAULT_ALIGNMENT
                fmt['border'] = dxf.border if dxf.border else self._DEFAULT_BORDER
                fmt['font'] = dxf.font if dxf.font else self._DEFAULT_FONT
                fmt['protection'] = dxf.protection if dxf.protection else self._DEFAULT_PROTECTION
                fmt['fill'] = PatternFill(fill_type=dxf.fill.fill_type, bgColor=dxf.fill.fgColor, fgColor=dxf.fill.bgColor) if dxf.fill else self._DEFAULT_FILL # NOTE: fgcolor and bgcolor are inversed in DifferentialStyle

            # Read formula
            if column.calculatedColumnFormula:
                formula = column.calculatedColumnFormula
                if formula.array:
                    fmt['formula'] = ArrayFormula(formula.attr_text)
                else:
                    fmt['formula'] = '=' + formula.attr_text

                #logger.debug('column %s (%s) formula: %s', get_column_letter(index + 1), column.name, fmt['formula'])
            
        return fmt_list


    _DEFAULT_NUMBER_FORMAT = 'General'
    _DEFAULT_FILL = PatternFill(fill_type=None)
    _DEFAULT_ALIGNMENT = None # openpyxl.styles.alignment.Alignment
    _DEFAULT_BORDER = None # openpyxl.styles.alignment.Border
    _DEFAULT_FONT = None # openpyxl.styles.fonts.Font
    _DEFAULT_PROTECTION = None # openpyxl.styles.protection.Protection


    def __iter__(self):
        return self.Iterator(self)
    
    def iterate(self, readonly: bool = False):
        return self.Iterator(self, readonly=readonly)

    class Iterator:
        def __init__(self, table: ExcelTable, readonly: bool = False):
            self.next_index = 0
            self.table = table
            self.readonly = readonly

        def __next__(self):
            if self.next_index >= self.table.row_count:
                raise StopIteration()
            
            row = self.table.get_row(self.next_index, readonly=self.readonly)
            self.next_index += 1
            return row
